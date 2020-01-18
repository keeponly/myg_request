
from openpyxl import load_workbook
from test_common import project_path
from test_common.read_config import ReadConfig


class DoExcel:
    """该类完成测试数据的读取，以及测试结果的写回"""
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self, section): # 配置文件里面的片段名
        """读取excel文件"""
        # 从配置文件获取读取那些测试数据
        case_id = ReadConfig(project_path.conf_path).get_data(section, 'case_id')
        wb = load_workbook(self.file_name)  # 打开工作簿
        sheet = wb[self.sheet_name]  # 定位表单
        tel =self.get_tel()
        test_data = []
        for i in range(2, sheet.max_row+1):
            #print(i)
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value


            if sheet.cell(i, 6).value.find('project6') != -1:  # 注意这个方法的使用以及返回值 也可以用成员运算符
                row_data['Params'] = sheet.cell(i, 6).value.replace('project6',"project_"+ str(tel))  # 替换值 tel
                self.update_tel(int(tel)+1)
            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()
        final_data = []
        if case_id =='all':
            final_data = test_data
        else:
            for i in case_id:
                final_data.append(test_data[i-1])
        return final_data
    def get_tel(self):
        """获取excel里面的手机号码"""
        wb = load_workbook(self.file_name)
        sheet = wb['tell']
        wb.close()
        return sheet.cell(1, 2).value

    def update_tel(self, new_tel):
        """写回手号码"""
        wb = load_workbook(self.file_name)
        sheet = wb['tell']
        sheet.cell(1, 2, new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self, row, col, value):
        """写回测试结果从0开始计算，所以加1"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value  # 函数中的参数value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    file_name = r'H:\Http_test\test_case\test_api.xlsx'
    sheet_name = 'add_loan'  # recharge
    res = DoExcel(file_name, sheet_name).read_excel('AddLOANCASE')
    print(res)

